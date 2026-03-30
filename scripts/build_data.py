from __future__ import annotations

import contextlib
import csv
import io
import json
import os
import pathlib
import re
import sys
import urllib.parse
import urllib.request
import zipfile
from dataclasses import dataclass

ROOT = pathlib.Path(__file__).resolve().parents[1]
RAW_DIR = ROOT / "raw"
DOWNLOADS_DIR = RAW_DIR / "downloads"
EXTRACT_DIR = RAW_DIR / "source_zip_3_6"
OUTPUT_DIR = ROOT / "data"

if (ROOT / ".deps").exists():
    sys.path.insert(0, str(ROOT / ".deps"))

from openpyxl import load_workbook  # type: ignore
import xlrd  # type: ignore

DATA_GO_FILES = [
    {
        "public_data_pk": "15048206",
        "public_data_detail_pk": "uddi:2ec04202-cc6f-49f3-b571-8d47d39c3c38",
        "filename": "제5회 전국동시지방선거 개표결과.xlsx",
        "election_round": 5,
        "election_date": "2010-06-02",
        "election_label": "제5회 전국동시지방선거",
    },
    {
        "public_data_pk": "15048207",
        "public_data_detail_pk": "uddi:ffd56e2e-4ae6-4557-a6e1-ef9d8c3401da",
        "filename": "제6회 전국동시지방선거 개표결과.xlsx",
        "election_round": 6,
        "election_date": "2014-06-04",
        "election_label": "제6회 전국동시지방선거",
    },
    {
        "public_data_pk": "15048208",
        "public_data_detail_pk": "uddi:fbe2a2e7-c7db-4950-bf59-aa8307d6b0ae",
        "filename": "제7회 전국동시지방선거 개표결과.xlsx",
        "election_round": 7,
        "election_date": "2018-06-13",
        "election_label": "제7회 전국동시지방선거",
    },
    {
        "public_data_pk": "15101509",
        "public_data_detail_pk": "uddi:6286249a-5c06-42a1-bf0f-d2fd0eb2f773",
        "filename": "제8회 전국동시지방선거 개표결과.xlsx",
        "election_round": 8,
        "election_date": "2022-06-01",
        "election_label": "제8회 전국동시지방선거",
    },
]

NEC_ARCHIVE = {
    "url": "https://www.nec.go.kr/common/board/Download.do?bcIdx=14979&cbIdx=1129&streFileNm=BBS_201808220327135370.zip",
    "filename": "nec_local_elections_3_to_6.zip",
}

SOURCES = [
    {
        "name": "중앙선거관리위원회 자료공간 - 전국동시지방선거 개표결과(제3회~제6회)",
        "url": "https://www.nec.go.kr/site/nec/ex/bbs/View.do?bcIdx=14979&cbIdx=1129",
    },
    {
        "name": "공공데이터포털 - 중앙선거관리위원회_제5회 전국동시지방선거 개표결과_20100602",
        "url": "https://www.data.go.kr/data/15048206/fileData.do",
    },
    {
        "name": "공공데이터포털 - 중앙선거관리위원회_제6회 전국동시지방선거 개표결과_20140604",
        "url": "https://www.data.go.kr/data/15048207/fileData.do",
    },
    {
        "name": "공공데이터포털 - 중앙선거관리위원회_제7회 전국동시지방선거 개표결과_20180613",
        "url": "https://www.data.go.kr/data/15048208/fileData.do",
    },
    {
        "name": "공공데이터포털 - 중앙선거관리위원회_제8회 전국동시지방선거 개표결과_20220601",
        "url": "https://www.data.go.kr/data/15101509/fileData.do",
    },
]


@dataclass
class TurnoutRecord:
    election_round: int
    election_label: str
    election_date: str
    province: str
    municipality: str
    municipality_key: str
    electorate: int
    votes: int
    invalid_votes: int
    abstentions: int

    @property
    def turnout_rate(self) -> float:
        return round((self.votes / self.electorate) * 100, 2) if self.electorate else 0.0

    def to_dict(self) -> dict[str, object]:
        return {
            "election_round": self.election_round,
            "election_label": self.election_label,
            "election_date": self.election_date,
            "province": self.province,
            "municipality": self.municipality,
            "municipality_key": self.municipality_key,
            "electorate": self.electorate,
            "votes": self.votes,
            "invalid_votes": self.invalid_votes,
            "abstentions": self.abstentions,
            "turnout_rate": self.turnout_rate,
        }


def parse_int(value: object) -> int:
    if value is None:
        return 0
    if isinstance(value, (int, float)):
        return int(value)
    text = str(value).strip()
    if not text:
        return 0
    text = re.sub(r"[^0-9-]", "", text)
    return int(text) if text else 0


def fetch_data_go_file(meta: dict[str, str]) -> pathlib.Path:
    DOWNLOADS_DIR.mkdir(parents=True, exist_ok=True)
    target = DOWNLOADS_DIR / meta["filename"]
    if target.exists():
        return target

    payload = urllib.parse.urlencode(
        {
            "publicDataPk": meta["public_data_pk"],
            "publicDataDetailPk": meta["public_data_detail_pk"],
        }
    ).encode()
    request = urllib.request.Request(
        "https://www.data.go.kr/tcs/dss/selectFileDataDownload.do?recommendDataYn=Y",
        data=payload,
        method="POST",
        headers={
            "User-Agent": "Mozilla/5.0",
            "Referer": f"https://www.data.go.kr/data/{meta['public_data_pk']}/fileData.do",
        },
    )
    response = json.loads(urllib.request.urlopen(request, timeout=30).read().decode("utf-8", "ignore"))
    file_url = (
        "https://www.data.go.kr/cmm/cmm/fileDownload.do"
        f"?atchFileId={response['atchFileId']}"
        f"&fileDetailSn={response['fileDetailSn']}"
        f"&dataNm={urllib.parse.quote(response['dataSetFileDetailInfo']['dataNm'])}"
    )
    target.write_bytes(urllib.request.urlopen(file_url, timeout=180).read())
    return target


def ensure_nec_archive() -> pathlib.Path:
    RAW_DIR.mkdir(parents=True, exist_ok=True)
    archive_path = RAW_DIR / NEC_ARCHIVE["filename"]
    if not archive_path.exists():
        archive_path.write_bytes(urllib.request.urlopen(NEC_ARCHIVE["url"], timeout=180).read())
    if not EXTRACT_DIR.exists():
        with zipfile.ZipFile(archive_path) as archive:
            archive.extractall(EXTRACT_DIR)
    return archive_path


def parse_xlsx_round(meta: dict[str, str]) -> list[TurnoutRecord]:
    workbook_path = fetch_data_go_file(meta)
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    sheet = workbook["시·도지사"]
    records: list[TurnoutRecord] = []

    for row in sheet.iter_rows(values_only=True):
        if meta["election_round"] == 5:
            province, municipality, town = row[0], row[1], row[2]
            electorate, votes, invalid_votes, abstentions = row[3], row[4], row[12], row[13]
            total_marker = "합계"
        elif meta["election_round"] == 6:
            province, municipality, town = row[0], row[1], row[2]
            electorate, votes, invalid_votes, abstentions = row[4], row[5], row[13], row[14]
            total_marker = "합계"
        elif meta["election_round"] == 7:
            province, municipality, town = row[2], row[3], row[4]
            electorate, votes, invalid_votes, abstentions = row[6], row[7], row[18], row[19]
            total_marker = "계"
        else:
            province, municipality, town = row[0], row[1], row[2]
            electorate, votes, invalid_votes, abstentions = row[4], row[5], row[13], row[14]
            total_marker = "합계"

        province = province.strip() if isinstance(province, str) else ""
        municipality = municipality.strip() if isinstance(municipality, str) else ""
        town = town.strip() if isinstance(town, str) else ""

        if not province or not municipality or town != total_marker:
            continue

        electorate_i = parse_int(electorate)
        votes_i = parse_int(votes)
        if not electorate_i or not votes_i:
            continue

        records.append(
            TurnoutRecord(
                election_round=meta["election_round"],
                election_label=meta["election_label"],
                election_date=meta["election_date"],
                province=province,
                municipality=municipality,
                municipality_key=f"{province} {municipality}",
                electorate=electorate_i,
                votes=votes_i,
                invalid_votes=parse_int(invalid_votes),
                abstentions=parse_int(abstentions),
            )
        )

    workbook.close()
    return records


def top_level_name(path: pathlib.Path, base: pathlib.Path) -> str:
    return path.parts[len(base.parts)]


def second_level_name(path: pathlib.Path, base: pathlib.Path) -> str:
    return path.parts[len(base.parts) + 1]


def office_to_province(office_title: str) -> str:
    clean = office_title.strip().strip("[]")
    replacements = [
        ("특별자치시장", "특별자치시"),
        ("특별자치도지사", "특별자치도"),
        ("특별시장", "특별시"),
        ("광역시장", "광역시"),
        ("도지사", "도"),
        ("시장", "시"),
    ]
    for suffix, target in replacements:
        if clean.endswith(suffix):
            return clean[: -len(suffix)] + target
    return clean


def decode_legacy_text(value: object) -> str:
    text = str(value).strip()
    if not text:
        return ""
    try:
        return text.encode("latin1").decode("cp949").strip()
    except (UnicodeEncodeError, UnicodeDecodeError):
        return text


def decode_legacy_zip_name(name: str) -> str:
    try:
        return name.encode("cp437").decode("cp949")
    except (UnicodeEncodeError, UnicodeDecodeError):
        return name


def normalize_2002_province(name: str) -> str:
    mapping = {
        "서울": "서울특별시",
        "부산": "부산광역시",
        "대구": "대구광역시",
        "인천": "인천광역시",
        "광주": "광주광역시",
        "대전": "대전광역시",
        "울산": "울산광역시",
        "경기": "경기도",
        "강원": "강원도",
        "충북": "충청북도",
        "충남": "충청남도",
        "전북": "전라북도",
        "전남": "전라남도",
        "경북": "경상북도",
        "경남": "경상남도",
        "제주": "제주도",
    }
    return mapping.get(name.strip(), name.strip())


def normalize_2006_municipality(name: str) -> str:
    clean = name.strip()
    if clean.endswith("시장"):
        return clean[:-2] + "시"
    if clean.endswith("군수"):
        return clean[:-2] + "군"
    clean = re.sub(r"^(.+?시)(.+구)$", r"\1 \2", clean)
    return re.sub(r"\s+", " ", clean).strip()


def normalize_2002_municipality(name: str) -> str:
    clean = decode_legacy_text(name)
    clean = re.sub(r"\([^)]*\)$", "", clean).strip()
    return normalize_2006_municipality(clean)


def parse_2002_round() -> list[TurnoutRecord]:
    ensure_nec_archive()
    records_by_key: dict[tuple[str, str], TurnoutRecord] = {}

    for path in EXTRACT_DIR.rglob("*.zip"):
        with zipfile.ZipFile(path) as archive:
            members = [
                (member_name, decode_legacy_zip_name(member_name))
                for member_name in archive.namelist()
                if member_name.lower().endswith(".xls")
            ]

            if len(members) != 16:
                continue

            for member_name, decoded_name in members:
                province = normalize_2002_province(pathlib.Path(decoded_name).stem)
                with archive.open(member_name) as member_stream:
                    workbook = xlrd.open_workbook(
                        file_contents=member_stream.read(),
                        on_demand=True,
                        logfile=open(os.devnull, "w"),
                    )
                sheet = workbook.sheet_by_index(0)
                header_name = decode_legacy_text(sheet.cell_value(0, 0))
                subheader_name = decode_legacy_text(sheet.cell_value(0, 1))
                if header_name != "위원회명" or subheader_name != "투표구명":
                    workbook.release_resources()
                    continue

                header_row = [decode_legacy_text(cell) for cell in sheet.row_values(2)]
                valid_total_idx = next((idx for idx, value in enumerate(header_row) if value == "계"), None)

                for row_idx in range(3, sheet.nrows):
                    municipality_raw = decode_legacy_text(sheet.cell_value(row_idx, 0))
                    label = decode_legacy_text(sheet.cell_value(row_idx, 1))
                    electorate = parse_int(sheet.cell_value(row_idx, 2))
                    votes = parse_int(sheet.cell_value(row_idx, 3))
                    if not municipality_raw or label != "합계" or not electorate or not votes:
                        continue

                    municipality = normalize_2002_municipality(municipality_raw)
                    valid_votes = parse_int(sheet.cell_value(row_idx, valid_total_idx)) if valid_total_idx is not None else 0
                    record = TurnoutRecord(
                        election_round=3,
                        election_label="제3회 전국동시지방선거",
                        election_date="2002-06-13",
                        province=province,
                        municipality=municipality,
                        municipality_key=f"{province} {municipality}",
                        electorate=electorate,
                        votes=votes,
                        invalid_votes=max(votes - valid_votes, 0) if valid_votes else 0,
                        abstentions=max(electorate - votes, 0),
                    )
                    records_by_key[(province, municipality)] = record

                workbook.release_resources()

    records = list(records_by_key.values())
    records.sort(key=lambda item: (item.province, item.municipality))
    return records


def parse_2006_round() -> list[TurnoutRecord]:
    ensure_nec_archive()
    records: list[TurnoutRecord] = []

    files = [
        path
        for path in EXTRACT_DIR.rglob("*.xls")
        if "4" in top_level_name(path, EXTRACT_DIR) and second_level_name(path, EXTRACT_DIR).startswith("1_")
    ]

    for path in files:
        workbook = xlrd.open_workbook(str(path), on_demand=True, logfile=open(os.devnull, "w"))
        sheet = workbook.sheet_by_index(0)
        if sheet.nrows < 8 or sheet.ncols < 17:
            workbook.release_resources()
            continue

        office = str(sheet.cell_value(1, 1)).strip()
        municipality = ""
        for cell in sheet.row_values(1):
            cell_text = str(cell).strip()
            if cell_text.startswith("[") and cell_text.endswith("]") and cell_text != office:
                municipality = normalize_2006_municipality(cell_text.strip("[]"))
        summary_label = str(sheet.cell_value(6, 1)).strip()
        if not office or not municipality or summary_label != "합계":
            workbook.release_resources()
            continue

        electorate = parse_int(sheet.cell_value(6, 2))
        votes = parse_int(sheet.cell_value(6, 4))
        if not electorate or not votes:
            workbook.release_resources()
            continue

        province = office_to_province(office)
        header_row = [str(cell).strip() for cell in sheet.row_values(3)]
        invalid_idx = next((idx for idx, value in enumerate(header_row) if "무효" in value), 15)
        abstention_idx = next((idx for idx, value in enumerate(header_row) if "기권" in value), 16)
        records.append(
            TurnoutRecord(
                election_round=4,
                election_label="제4회 전국동시지방선거",
                election_date="2006-05-31",
                province=province,
                municipality=municipality,
                municipality_key=f"{province} {municipality}",
                electorate=electorate,
                votes=votes,
                invalid_votes=parse_int(sheet.cell_value(6, invalid_idx)),
                abstentions=parse_int(sheet.cell_value(6, abstention_idx)),
            )
        )
        workbook.release_resources()

    records.sort(key=lambda item: (item.province, item.municipality))
    return records


def parse_2006_round_stable() -> list[TurnoutRecord]:
    ensure_nec_archive()
    records: list[TurnoutRecord] = []
    files = [
        path
        for path in EXTRACT_DIR.rglob("*.xls")
        if "4" in top_level_name(path, EXTRACT_DIR) and second_level_name(path, EXTRACT_DIR).startswith("1_")
    ]

    for path in files:
        workbook = xlrd.open_workbook(str(path), on_demand=True, logfile=open(os.devnull, "w"))
        sheet = workbook.sheet_by_index(0)
        office = str(sheet.cell_value(1, 1)).strip()
        municipality = ""
        for cell in sheet.row_values(1):
            cell_text = str(cell).strip()
            if cell_text.startswith("[") and cell_text.endswith("]") and cell_text != office:
                municipality = normalize_2006_municipality(cell_text.strip("[]"))

        summary_label = str(sheet.cell_value(6, 1)).strip()
        electorate = parse_int(sheet.cell_value(6, 2))
        votes = parse_int(sheet.cell_value(6, 4))
        if office and municipality and summary_label == "합계" and electorate and votes:
            province = office_to_province(office)
            header_row = [str(cell).strip() for cell in sheet.row_values(3)]
            invalid_idx = next((idx for idx, value in enumerate(header_row) if "무효" in value), 15)
            abstention_idx = next((idx for idx, value in enumerate(header_row) if "기권" in value), 16)
            records.append(
                TurnoutRecord(
                    election_round=4,
                    election_label="제4회 전국동시지방선거",
                    election_date="2006-05-31",
                    province=province,
                    municipality=municipality,
                    municipality_key=f"{province} {municipality}",
                    electorate=electorate,
                    votes=votes,
                    invalid_votes=parse_int(sheet.cell_value(6, invalid_idx)),
                    abstentions=parse_int(sheet.cell_value(6, abstention_idx)),
                )
            )
        workbook.release_resources()

    records.sort(key=lambda item: (item.province, item.municipality))
    return records


def build_records() -> list[TurnoutRecord]:
    records: list[TurnoutRecord] = []
    records.extend(parse_2002_round())
    records.extend(parse_2006_round_stable())
    for meta in DATA_GO_FILES:
        records.extend(parse_xlsx_round(meta))
    records.sort(key=lambda item: (item.election_date, item.province, item.municipality))
    return records


def write_outputs(records: list[TurnoutRecord]) -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    csv_path = OUTPUT_DIR / "local-election-turnout-municipal.csv"
    json_path = OUTPUT_DIR / "local-election-turnout-municipal.json"
    sources_path = OUTPUT_DIR / "sources.json"

    fieldnames = [
        "election_round",
        "election_label",
        "election_date",
        "province",
        "municipality",
        "municipality_key",
        "electorate",
        "votes",
        "invalid_votes",
        "abstentions",
        "turnout_rate",
    ]

    with csv_path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for record in records:
            writer.writerow(record.to_dict())

    payload = {
        "coverage": {
            "from": records[0].election_date if records else None,
            "to": records[-1].election_date if records else None,
            "election_rounds": sorted({record.election_round for record in records}),
            "municipality_count": len({record.municipality_key for record in records}),
            "record_count": len(records),
        },
        "records": [record.to_dict() for record in records],
    }

    json_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    sources_path.write_text(json.dumps(SOURCES, ensure_ascii=False, indent=2), encoding="utf-8")


def main() -> None:
    records = build_records()
    write_outputs(records)
    print(f"wrote {len(records)} records")


if __name__ == "__main__":
    main()
